# -*- coding: utf-8 -*-
"""
Schritt 2+3 des A-Team-Plans: Classifier trainieren & fair evaluieren.

Vergleicht ZWEI Ansätze auf DEMSELBEN Datensatz / denselben Folds:
  (A) Regex-Baseline  -> Keyword-/Regex-Matching auf der Caption ("Regex A")
  (B) Classifier      -> TF-IDF + One-vs-Rest Logistic Regression

Problem ist MULTI-LABEL: eine Abbildung kann gleichzeitig Plot, Diagram, Photo sein.
Deshalb 3 unabhängige binäre Labels (plot / diagram / photo) aus den manuell
annotierten Spalten manual_plot / manual_diagram / manual_photo.

Bewertung über k-fold Cross-Validation, damit nie auf Trainingsdaten gemessen wird.
Berichtet Precision / Recall / F1 pro Klasse, Macro-F1 und Exact-Match (Subset-Accuracy).

Benötigt:  pip install scikit-learn pandas
Aufruf:    python train_and_evaluate.py  [pfad/zur/annotation.csv]

Läuft auch auf TEILDATEN: es nutzt einfach alle Zeilen, die schon ein
manuelles Label haben (Skip / Needs Review werden ausgeschlossen).
"""

import os
import re
import sys
import numpy as np
import pandas as pd

# Windows-Konsole UTF-8-sicher machen (verhindert Crash bei Umlauten/Sonderzeichen)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.multiclass import OneVsRestClassifier
from sklearn.pipeline import Pipeline
from sklearn.model_selection import KFold, GroupKFold, cross_val_predict
from sklearn.metrics import (classification_report, f1_score, accuracy_score,
                             precision_recall_fscore_support)


# ======================================================================
#  Konfiguration
# ======================================================================
DEFAULT_CSV = "annotation_task_1000.csv"
TEXT_COL    = "caption"
CLASSES     = ["plot", "diagram", "photo"]          # Reihenfolge der binären Labels
FLAG_COLS   = {"plot": "manual_plot",
               "diagram": "manual_diagram",
               "photo": "manual_photo"}
UNCLEAR_COL = "unclear"
LABEL_COL   = "manual_label"                          # vom Annotations-Tool gesetzt
N_FOLDS     = 5
RANDOM_STATE = 42

# "Regex B" — Schlüsselwörter aus den Kategorie-Definitionen von Lee et al.
# Zentrale Quelle: regex_lee.py (muss im selben Ordner liegen).
from regex_lee import regex_predict, REGEX_KEYWORDS


# ======================================================================
#  Daten laden & filtern
# ======================================================================
def load_annotated(csv_path):
    df = pd.read_csv(csv_path, sep=";", encoding="utf-8-sig", dtype=str).fillna("")

    def flag(v):
        return 1 if str(v).strip() in ("1", "1.0", "x", "X", "true", "True") else 0

    # binäre Zielspalten erzeugen
    for cls, col in FLAG_COLS.items():
        if col not in df.columns:
            raise SystemExit(f"Spalte '{col}' fehlt in der CSV.")
        df[f"y_{cls}"] = df[col].map(flag)

    has_unclear = df[UNCLEAR_COL].map(flag) if UNCLEAR_COL in df.columns else 0
    label_txt = df[LABEL_COL].str.strip() if LABEL_COL in df.columns else ""

    # Eine Zeile gilt als "annotiert & nutzbar", wenn mindestens ein
    # Klassen-Flag gesetzt ist (also nicht Skip / Needs Review / leer).
    y_any = df[[f"y_{c}" for c in CLASSES]].sum(axis=1) > 0
    usable = y_any & (df[TEXT_COL].str.strip() != "")

    n_total   = len(df)
    n_unclear = int((has_unclear == 1).sum()) if hasattr(has_unclear, "sum") else 0
    n_skip = 0
    if LABEL_COL in df.columns:
        n_skip = int((label_txt.str.lower() == "skip").sum())

    train = df[usable].copy()
    return df, train, dict(n_total=n_total, n_usable=len(train),
                           n_unclear=n_unclear, n_skip=n_skip)


def get_XY(df):
    X = df[TEXT_COL].astype(str).values
    Y = df[[f"y_{c}" for c in CLASSES]].values.astype(int)
    return X, Y


# ======================================================================
#  Regex-Baseline  (regex_predict wird aus regex_lee importiert, s. o.)
# ======================================================================
def majority_predict(Y_ref, n_rows):
    """Triviale Baseline: sagt fuer jede Klasse immer den haeufigsten Wert vorher
    (bei euch praktisch: immer 'Plot', nie Diagram/Photo). Lernt nichts und dient
    nur als unterster Vergleichsmassstab - alles muss darueber liegen."""
    maj = (Y_ref.mean(axis=0) >= 0.5).astype(int)
    return np.tile(maj, (n_rows, 1))


# ======================================================================
#  Classifier
# ======================================================================
def build_classifier():
    return Pipeline([
        ("tfidf", TfidfVectorizer(ngram_range=(1, 2), min_df=2,
                                  sublinear_tf=True, stop_words="english")),
        ("clf", OneVsRestClassifier(
            LogisticRegression(max_iter=2000, class_weight="balanced"))),
    ])


# ======================================================================
#  Reporting
# ======================================================================
def report(name, y_true, y_pred):
    print(f"\n{'='*64}\n  {name}\n{'='*64}")
    print(classification_report(y_true, y_pred, target_names=CLASSES,
                                zero_division=0, digits=3))
    macro = f1_score(y_true, y_pred, average="macro", zero_division=0)
    micro = f1_score(y_true, y_pred, average="micro", zero_division=0)
    exact = accuracy_score(y_true, y_pred)     # alle 3 Labels exakt richtig
    print(f"  Macro-F1 : {macro:.3f}")
    print(f"  Micro-F1 : {micro:.3f}")
    print(f"  Exact-Match (alle 3 Labels korrekt): {exact:.3f}")
    return dict(name=name, macro_f1=macro, micro_f1=micro, exact_match=exact)


def per_class_f1(y_true, y_pred):
    return {c: f1_score(y_true[:, j], y_pred[:, j], zero_division=0)
            for j, c in enumerate(CLASSES)}


# ======================================================================
#  Hauptablauf
# ======================================================================
def _scores_tables(Y, preds_dict):
    """Erzeugt zwei Tabellen: Scores pro Klasse + Gesamtwerte (je Methode)."""
    per_rows, agg_rows = [], []
    for name, pred in preds_dict.items():
        p, r, f, sup = precision_recall_fscore_support(
            Y, pred, average=None, zero_division=0)
        for j, c in enumerate(CLASSES):
            per_rows.append({"Methode": name, "Klasse": c,
                             "Precision": round(float(p[j]), 3),
                             "Recall": round(float(r[j]), 3),
                             "F1": round(float(f[j]), 3),
                             "Anzahl": int(sup[j])})
        agg_rows.append({"Methode": name,
                         "Macro-F1": round(f1_score(Y, pred, average="macro", zero_division=0), 3),
                         "Micro-F1": round(f1_score(Y, pred, average="micro", zero_division=0), 3),
                         "Weighted-F1": round(f1_score(Y, pred, average="weighted", zero_division=0), 3),
                         "Exact-Match": round(accuracy_score(Y, pred), 3)})
    return pd.DataFrame(per_rows), pd.DataFrame(agg_rows)


def top_words_table(X, Y, top_n=30):
    """Linguistische Signatur: die top_n trennschaerfsten Begriffe je Klasse
    (TF-IDF + Logistic Regression). Beschreibt, welche Woerter das Modell pro
    Klasse am staerksten gewichtet."""
    vec = TfidfVectorizer(ngram_range=(1, 2), min_df=5, sublinear_tf=True,
                          stop_words="english")
    Xt = vec.fit_transform(X)
    feats = np.array(vec.get_feature_names_out())
    rows = []
    for j, c in enumerate(CLASSES):
        if Y[:, j].sum() == 0:
            continue
        clf = LogisticRegression(max_iter=2000, class_weight="balanced").fit(Xt, Y[:, j])
        order = np.argsort(clf.coef_[0])[::-1][:top_n]
        for rank, i in enumerate(order, 1):
            rows.append({"Klasse": c, "Rang": rank, "Begriff": feats[i],
                         "Gewicht": round(float(clf.coef_[0][i]), 3)})
    return pd.DataFrame(rows)


def write_scores_xlsx(csv_path, Y, preds_dict, details_df, out_path=None, signature_df=None):
    """Excel mit Blatt 'Scores', 'Details' und optional 'Signatur' (Top-Woerter)."""
    import openpyxl
    from openpyxl.styles import Font
    per_df, agg_df = _scores_tables(Y, preds_dict)
    xlsx_path = out_path or (os.path.splitext(csv_path)[0] + "_evaluation.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"
    bold = Font(bold=True)
    title = Font(bold=True, size=12)

    r = 1
    ws.cell(r, 1, "Scores pro Klasse").font = title; r += 2
    cols1 = list(per_df.columns)
    for j, c in enumerate(cols1, 1):
        ws.cell(r, j, c).font = bold
    r += 1
    for _, row in per_df.iterrows():
        for j, c in enumerate(cols1, 1):
            ws.cell(r, j, row[c])
        r += 1
    r += 1
    ws.cell(r, 1, "Gesamtwerte").font = title; r += 2
    cols2 = list(agg_df.columns)
    for j, c in enumerate(cols2, 1):
        ws.cell(r, j, c).font = bold
    r += 1
    for _, row in agg_df.iterrows():
        for j, c in enumerate(cols2, 1):
            ws.cell(r, j, row[c])
        r += 1
    for col, w in zip("ABCDEFG", [12, 11, 11, 9, 8, 9, 12]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Details")
    ws2.append(list(details_df.columns))
    for cell in ws2[1]:
        cell.font = bold
    for _, row in details_df.iterrows():
        ws2.append([row[c] for c in details_df.columns])

    # Blatt 'Signatur' = staerkste Woerter pro Klasse (linguistische Signatur)
    if signature_df is not None and len(signature_df):
        ws3 = wb.create_sheet("Signatur")
        ws3.cell(1, 1, "Staerkste Woerter pro Klasse (TF-IDF + LogReg-Gewicht)").font = title
        cols = list(signature_df.columns)
        for j, c in enumerate(cols, 1):
            ws3.cell(3, j, c).font = bold
        r3 = 4
        for _, row in signature_df.iterrows():
            for j, c in enumerate(cols, 1):
                ws3.cell(r3, j, row[c])
            r3 += 1
        for col, w in zip("ABCD", [12, 7, 22, 10]):
            ws3.column_dimensions[col].width = w

    wb.save(xlsx_path)
    return xlsx_path


def pick_csv():
    """Nimmt den Pfad aus der Kommandozeile, sonst Datei-Dialog (wie annotate.py)."""
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        return sys.argv[1]
    if os.path.isfile(DEFAULT_CSV):
        return DEFAULT_CSV
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.withdraw()
    path = filedialog.askopenfilename(
        title="Annotierte CSV wählen",
        filetypes=[("CSV", "*.csv"), ("Alle Dateien", "*.*")])
    root.destroy()
    return path or None


def main():
    csv_path = pick_csv()
    if not csv_path:
        print("Keine CSV gewählt. Abbruch.")
        return

    df_all, df, stats = load_annotated(csv_path)
    print(f"Datei: {csv_path}")
    print(f"Zeilen gesamt          : {stats['n_total']}")
    print(f"Nutzbar (mit Label)    : {stats['n_usable']}")
    print(f"davon Needs-Review     : {stats['n_unclear']}  (aus Training ausgeschlossen)")
    print(f"Skip                   : {stats['n_skip']}  (aus Training ausgeschlossen)")

    if stats["n_usable"] < 30:
        print("\n⚠ Noch zu wenige annotierte Zeilen für eine belastbare Evaluation.")
        print("  Mach erst weiter mit der Annotation – das Skript läuft dann automatisch durch.")
        # Verteilung trotzdem zeigen
        X, Y = get_XY(df)
        if len(Y):
            print("\nKlassenverteilung (annotiert):")
            for j, c in enumerate(CLASSES):
                print(f"  {c:8}: {int(Y[:, j].sum())}")
        return

    X, Y = get_XY(df)

    print("\nKlassenverteilung (positive Beispiele):")
    for j, c in enumerate(CLASSES):
        print(f"  {c:8}: {int(Y[:, j].sum())} / {len(Y)}")

    # Cross-Validation: wenn moeglich nach PAPER gruppieren (GroupKFold), damit nie
    # Figuren desselben Preprints gleichzeitig in Training und Test landen
    # -> verhindert Leakage auf Preprint-Ebene. Sonst normale KFold.
    groups = df["source_json"].values if "source_json" in df.columns else None
    if groups is not None and len(set(groups)) >= N_FOLDS:
        cv = GroupKFold(n_splits=N_FOLDS)
        cv_name = f"{N_FOLDS}-fold GroupKFold (Split nach Preprint, leakage-frei)"
    else:
        cv = KFold(n_splits=N_FOLDS, shuffle=True, random_state=RANDOM_STATE)
        groups = None
        cv_name = f"{N_FOLDS}-fold KFold"
    print(f"\nCross-Validation: {cv_name}")

    # (0) Triviale Baseline – immer die haeufigste Klasse raten (lernt nichts)
    y_dummy = majority_predict(Y, len(Y))

    # (A) Regex – braucht kein Training, direkt auf allen Zeilen vorhersagen
    y_regex = regex_predict(X)

    # (B) Classifier – ehrliche Out-of-Fold-Vorhersagen via Cross-Validation
    clf = build_classifier()
    y_clf = cross_val_predict(clf, X, Y, cv=cv, groups=groups, method="predict")

    r_dummy = report("Majority-Baseline (immer haeufigste Klasse)", Y, y_dummy)
    r_regex = report("Regex-Baseline (Regex A)", Y, y_regex)
    r_clf   = report(f"Classifier (TF-IDF + OvR LogReg, {cv_name})", Y, y_clf)

    # Direktvergleich pro Klasse: Dummy < Regex < Classifier
    f_dummy = per_class_f1(Y, y_dummy)
    f_regex = per_class_f1(Y, y_regex)
    f_clf   = per_class_f1(Y, y_clf)
    print(f"\n{'='*64}\n  Vergleich F1 pro Klasse\n{'='*64}")
    print(f"  {'Klasse':10} {'Dummy':>8} {'Regex':>8} {'Classifier':>12}")
    for c in CLASSES:
        print(f"  {c:10} {f_dummy[c]:>8.3f} {f_regex[c]:>8.3f} {f_clf[c]:>12.3f}")
    print(f"  {'MACRO':10} {r_dummy['macro_f1']:>8.3f} {r_regex['macro_f1']:>8.3f} "
          f"{r_clf['macro_f1']:>12.3f}")

    # Ergebnisse + Fehlerfälle speichern (für Teil 3 der Präsi)
    out = df.copy()
    for j, c in enumerate(CLASSES):
        out[f"pred_regex_{c}"] = y_regex[:, j]
        out[f"pred_clf_{c}"]   = y_clf[:, j]
        out[f"err_clf_{c}"]    = (y_clf[:, j] != Y[:, j]).astype(int)
    out_path = os.path.splitext(csv_path)[0] + "_eval_results.csv"
    out.to_csv(out_path, sep=";", index=False, encoding="utf-8-sig")
    print(f"\nDetailergebnisse (inkl. Fehlerfälle) gespeichert: {out_path}")
    print("Tipp: nach err_clf_* filtern, um typische Classifier-Fehler anzusehen.")

    # --- Ergebnis-Blatt (Excel): Gesamt-Scores + Details + Signatur ---
    import time
    preds = {"Majority": y_dummy, "Regex A": y_regex, "Classifier": y_clf}
    signature_df = top_words_table(X, Y, top_n=30)
    try:
        try:
            xlsx_path = write_scores_xlsx(csv_path, Y, preds, out, signature_df=signature_df)
        except PermissionError:
            alt = os.path.splitext(csv_path)[0] + f"_evaluation_{time.strftime('%H%M%S')}.xlsx"
            xlsx_path = write_scores_xlsx(csv_path, Y, preds, out, out_path=alt,
                                          signature_df=signature_df)
            print("Hinweis: Die Datei '..._evaluation.xlsx' war gesperrt "
                  "(vermutlich noch in Excel geöffnet) – stattdessen neue Datei geschrieben.")
        print(f"Ergebnis-Blatt (Scores + Details + Signatur) gespeichert: {xlsx_path}")
    except ImportError:
        sc_per, _ = _scores_tables(Y, preds)
        sc_path = os.path.splitext(csv_path)[0] + "_scores.csv"
        sc_per.to_csv(sc_path, sep=";", index=False, encoding="utf-8-sig")
        print(f"(openpyxl fehlt -> Scores als CSV gespeichert: {sc_path}; "
              f"für das Excel-Blatt einmalig: pip install openpyxl)")


if __name__ == "__main__":
    main()
