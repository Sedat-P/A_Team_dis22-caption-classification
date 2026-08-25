# -*- coding: utf-8 -*-
"""
Unabhaengige Evaluation gegen das selbst gelabelte Gold-Set.

Ablauf:
  1) Classifier wird auf der VOLLEN Trainings-CSV (z. B. annotation_Z.csv) trainiert.
  2) Er wird auf das separate, selbst gelabelte Eval-Set (annotation_task_eval.csv)
     angewendet - diese Figuren waren NICHT im Training.
  3) Die Vorhersagen werden mit deinen Gold-Labels verglichen -> Precision/Recall/F1.
  4) Zusaetzlich die Regex-Baseline auf demselben Eval-Set, zum Vergleich.

Das ist die ehrlichste Zahl: gemessen auf voellig unabhaengigen, von Hand gelabelten Daten.

Erzeugt:  <eval>_goldeval.xlsx  (Blatt 'Scores' + 'Details')
          <eval>_goldeval.csv   (jede Figur mit Gold-Label, Vorhersage, Fehler-Flag)

Benoetigt:  pip install scikit-learn pandas openpyxl
Aufruf:     python evaluate_on_goldset.py [train.csv] [eval.csv]
            (ohne Argumente: zwei Datei-Dialoge)

WICHTIG: 'train_and_evaluate.py' muss im selben Ordner liegen (wird importiert).
"""

import os
import sys
import time
import numpy as np
import pandas as pd

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import train_and_evaluate as te   # gleiche Pipeline/Config wiederverwenden
except ImportError:
    raise SystemExit("train_and_evaluate.py nicht gefunden - bitte in denselben Ordner legen.")


def pick(title):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.withdraw()
    p = filedialog.askopenfilename(title=title,
                                   filetypes=[("CSV", "*.csv"), ("Alle", "*.*")])
    root.destroy()
    return p or None


def error_analysis(X, Y, yp, classes):
    """Liefert drei Tabellen fuer die Fehleranalyse:
    - Fehler pro Klasse (TP/FP/FN + Recall/Precision)
    - Verwechslungs-Matrix (nur Einzel-Label-Faelle)
    - Beispiel-Captions je typischem Fehlertyp."""
    per = []
    for j, c in enumerate(classes):
        t = Y[:, j]; p = yp[:, j]
        tp = int(((p == 1) & (t == 1)).sum())
        fp = int(((p == 1) & (t == 0)).sum())
        fn = int(((p == 0) & (t == 1)).sum())
        rec = round(tp / (tp + fn), 3) if (tp + fn) else 0.0
        prec = round(tp / (tp + fp), 3) if (tp + fp) else 0.0
        per.append({"Klasse": c, "TP": tp, "FP (faelschlich)": fp,
                    "FN (uebersehen)": fn, "Recall": rec, "Precision": prec})
    per_df = pd.DataFrame(per)

    single = (Y.sum(1) == 1) & (yp.sum(1) == 1)
    cm = pd.DataFrame(0, index=classes, columns=classes)
    if single.sum() > 0:
        t1 = Y[single].argmax(1); p1 = yp[single].argmax(1)
        for a, b in zip(t1, p1):
            cm.iloc[a, b] += 1

    ex = []

    def add(mask, label, n=3):
        for i in np.where(mask)[0][:n]:
            ex.append({"Fehlertyp": label, "Caption": str(X[i])[:220]})

    idx = {c: j for j, c in enumerate(classes)}
    if {"plot", "diagram", "photo"} <= set(idx):
        pl, di, ph = idx["plot"], idx["diagram"], idx["photo"]
        add((Y[:, di] == 1) & (yp[:, di] == 0) & (yp[:, pl] == 1), "echt Diagram -> als Plot gelesen")
        add((Y[:, pl] == 1) & (yp[:, pl] == 0) & (yp[:, di] == 1), "echt Plot -> als Diagram gelesen")
        add((Y[:, ph] == 1) & (yp[:, ph] == 0), "echt Photo -> uebersehen")
        add((Y[:, ph] == 0) & (yp[:, ph] == 1), "faelschlich als Photo erkannt")
    ex_df = pd.DataFrame(ex)
    return per_df, cm, ex_df


def signature_from_pipeline(clf, classes, top_n=30):
    """Top-Woerter pro Klasse aus dem TRAINIERTEN Modell (genau dem bewerteten).
    Liest die Logistic-Regression-Gewichte je Klasse aus der gefitteten Pipeline."""
    vec = clf.named_steps["tfidf"]
    ovr = clf.named_steps["clf"]
    feats = np.array(vec.get_feature_names_out())
    rows = []
    for j, c in enumerate(classes):
        coef = ovr.estimators_[j].coef_[0]
        order = np.argsort(coef)[::-1][:top_n]
        for rank, i in enumerate(order, 1):
            rows.append({"Klasse": c, "Rang": rank, "Begriff": feats[i],
                         "Gewicht": round(float(coef[i]), 3)})
    return pd.DataFrame(rows)


def write_eval_xlsx(out_path, Y, preds_dict, details_df, err=None, signature_df=None):
    """Schreibt Excel mit Blatt 'Scores' + 'Details' (unabhaengig von der
    train_and_evaluate-Version, braucht kein out_path dort)."""
    import openpyxl
    from openpyxl.styles import Font
    per_df, agg_df = te._scores_tables(Y, preds_dict)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"
    bold = Font(bold=True)
    title = Font(bold=True, size=12)
    r = 1
    ws.cell(r, 1, "Scores pro Klasse").font = title; r += 2
    cols1 = list(per_df.columns)
    for j, c in enumerate(cols1, 1):
        ws.cell(r, j, c).font = bold
    r += 1
    for _, row in per_df.iterrows():
        for j, c in enumerate(cols1, 1):
            ws.cell(r, j, row[c])
        r += 1
    r += 1
    ws.cell(r, 1, "Gesamtwerte").font = title; r += 2
    cols2 = list(agg_df.columns)
    for j, c in enumerate(cols2, 1):
        ws.cell(r, j, c).font = bold
    r += 1
    for _, row in agg_df.iterrows():
        for j, c in enumerate(cols2, 1):
            ws.cell(r, j, row[c])
        r += 1
    for col, w in zip("ABCDEFG", [12, 11, 11, 9, 8, 9, 12]):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Details")
    ws2.append(list(details_df.columns))
    for cell in ws2[1]:
        cell.font = bold
    for _, row in details_df.iterrows():
        ws2.append([row[c] for c in details_df.columns])

    # Blatt 'Fehleranalyse'
    if err is not None:
        per_df, cm, ex_df = err
        wf = wb.create_sheet("Fehleranalyse")
        rr = 1
        wf.cell(rr, 1, "Fehler pro Klasse").font = title; rr += 2
        cols = list(per_df.columns)
        for j, c in enumerate(cols, 1):
            wf.cell(rr, j, c).font = bold
        rr += 1
        for _, row in per_df.iterrows():
            for j, c in enumerate(cols, 1):
                wf.cell(rr, j, row[c])
            rr += 1
        rr += 1
        wf.cell(rr, 1, "Verwechslungen (nur Einzel-Label-Faelle)").font = title; rr += 2
        wf.cell(rr, 1, "ECHT \\ VORHERGESAGT").font = bold
        for j, c in enumerate(cm.columns, 2):
            wf.cell(rr, j, c).font = bold
        rr += 1
        for c in cm.index:
            wf.cell(rr, 1, c).font = bold
            for j, cc in enumerate(cm.columns, 2):
                wf.cell(rr, j, int(cm.loc[c, cc]))
            rr += 1
        rr += 1
        wf.cell(rr, 1, "Beispiel-Fehler").font = title; rr += 2
        if len(ex_df):
            for j, c in enumerate(ex_df.columns, 1):
                wf.cell(rr, j, c).font = bold
            rr += 1
            for _, row in ex_df.iterrows():
                for j, c in enumerate(ex_df.columns, 1):
                    wf.cell(rr, j, row[c])
                rr += 1
        wf.column_dimensions["A"].width = 32
        wf.column_dimensions["B"].width = 70

    # Blatt 'Signatur' = staerkste Woerter pro Klasse (aus dem trainierten Modell)
    if signature_df is not None and len(signature_df):
        wsig = wb.create_sheet("Signatur")
        wsig.cell(1, 1, "Staerkste Woerter pro Klasse (Gewicht des trainierten Modells)").font = title
        cols = list(signature_df.columns)
        for j, c in enumerate(cols, 1):
            wsig.cell(3, j, c).font = bold
        rs = 4
        for _, row in signature_df.iterrows():
            for j, c in enumerate(cols, 1):
                wsig.cell(rs, j, row[c])
            rs += 1
        for col, w in zip("ABCD", [12, 7, 22, 10]):
            wsig.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path


def main():
    train_csv = sys.argv[1] if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]) else None
    eval_csv = sys.argv[2] if len(sys.argv) > 2 and os.path.isfile(sys.argv[2]) else None
    if not train_csv:
        train_csv = pick("TRAININGS-CSV waehlen (z. B. annotation_Z.csv)")
    if train_csv and not eval_csv:
        eval_csv = pick("EVAL-CSV waehlen (annotation_task_eval.csv, selbst gelabelt)")
    if not train_csv or not eval_csv:
        print("Train- oder Eval-CSV fehlt. Abbruch."); return

    # --- Daten laden ---
    _, train, st_tr = te.load_annotated(train_csv)
    _, ev, st_ev = te.load_annotated(eval_csv)
    print(f"Training: {st_tr['n_usable']} nutzbare Zeilen (von {st_tr['n_total']})")
    print(f"Eval    : {st_ev['n_usable']} nutzbare Zeilen (von {st_ev['n_total']})")
    if st_ev["n_usable"] == 0:
        raise SystemExit("Eval-Set hat keine gelabelten Zeilen - in annotate.py gelabelt?")

    X_tr, Y_tr = te.get_XY(train)
    X_ev, Y_ev = te.get_XY(ev)

    print("\nGold-Verteilung im Eval-Set:")
    for j, c in enumerate(te.CLASSES):
        print(f"  {c:8s}: {int(Y_ev[:, j].sum())} / {len(Y_ev)}")

    # --- Classifier auf vollem Training trainieren, auf Eval anwenden ---
    clf = te.build_classifier()
    clf.fit(X_tr, Y_tr)
    y_clf = clf.predict(X_ev)

    # --- Regex-Baseline + Majority-Baseline auf Eval ---
    y_regex = te.regex_predict(X_ev)
    y_dummy = te.majority_predict(Y_tr, len(Y_ev))

    # --- Berichte ---
    te.report("Regex-Baseline (Regex A) -- GOLD-EVAL", Y_ev, y_regex)
    te.report("Classifier (trainiert auf voller CSV) -- GOLD-EVAL", Y_ev, y_clf)

    print(f"\n{'='*64}\n  Vergleich F1 pro Klasse (Gold-Eval)\n{'='*64}")
    fr = te.per_class_f1(Y_ev, y_regex)
    fc = te.per_class_f1(Y_ev, y_clf)
    print(f"  {'Klasse':12s}{'Regex':>8s}{'Classifier':>13s}{'Diff':>9s}")
    for c in te.CLASSES:
        print(f"  {c:12s}{fr[c]:8.3f}{fc[c]:13.3f}{fc[c]-fr[c]:+9.3f}")

    # --- Fehleranalyse (auf dem unabhaengigen Eval-Set) ---
    err = error_analysis(X_ev, Y_ev, y_clf, te.CLASSES)
    per_df, cm, ex_df = err
    print(f"\n{'='*64}\n  Fehleranalyse (Classifier auf Eval)\n{'='*64}")
    print(per_df.to_string(index=False))
    print("\n  Verwechslungen (nur Einzel-Label) ECHT \\ VORHERGESAGT:")
    print(cm.to_string())

    # --- Details-Tabelle ---
    out = ev.copy().reset_index(drop=True)
    for j, c in enumerate(te.CLASSES):
        out[f"gold_{c}"] = Y_ev[:, j]
        out[f"pred_clf_{c}"] = y_clf[:, j]
        out[f"err_clf_{c}"] = (Y_ev[:, j] != y_clf[:, j]).astype(int)
    keep = ([te.TEXT_COL]
            + [f"gold_{c}" for c in te.CLASSES]
            + [f"pred_clf_{c}" for c in te.CLASSES]
            + [f"err_clf_{c}" for c in te.CLASSES])
    keep = [c for c in keep if c in out.columns]
    details = out[keep]

    base = os.path.splitext(eval_csv)[0] + "_goldeval"
    csv_out = base + ".csv"
    details.to_csv(csv_out, sep=";", index=False, encoding="utf-8-sig")
    print(f"\nDetails gespeichert: {csv_out}")

    # --- Signatur-Woerter aus genau diesem trainierten Modell ---
    signature_df = signature_from_pipeline(clf, te.CLASSES, top_n=30)

    # --- Excel (Scores + Details + Fehleranalyse + Signatur), mit Sperr-Fallback ---
    preds = {"Majority": y_dummy, "Regex A": y_regex, "Classifier": y_clf}
    try:
        try:
            xlsx = write_eval_xlsx(base + ".xlsx", Y_ev, preds, details,
                                   err=err, signature_df=signature_df)
        except PermissionError:
            xlsx = write_eval_xlsx(base + f"_{time.strftime('%H%M%S')}.xlsx",
                                   Y_ev, preds, details, err=err, signature_df=signature_df)
            print("Hinweis: Datei war gesperrt (in Excel offen?) - neue Datei geschrieben.")
        print(f"Ergebnis-Blatt (Scores + Details + Fehleranalyse + Signatur) gespeichert: {xlsx}")
    except ImportError:
        print("(openpyxl fehlt -> kein Excel-Blatt; einmalig: pip install openpyxl)")

    print("\nFertig. Die Zahlen im Blatt 'Scores' sind deine unabhaengige Evaluation.")


if __name__ == "__main__":
    main()
